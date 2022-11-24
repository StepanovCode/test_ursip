from copy import copy
from openpyxl import load_workbook

from app.services.formulars import FormularQ


class Worksheet_Form:

    def reformatting_sheet(self):
        raise NotImplementedError("В дочернем классе не переопределен метод unmerge_cells")

    def check_form_formular(self):
        raise NotImplementedError("В дочернем классе не переопределен метод unmerge_cells")


class WorksheetFormularQ(Worksheet_Form):
    def __init__(self, path: str,
                 height_header: int, max_column: int):
        self.path = path
        self.wb = None
        self.ws = None
        self.bounds = ()
        self.height_header = height_header
        self.max_column = max_column
        self.list_merge_cells = []
        self.copy_cells_ok = False
        self.merge_header_ok = False
        self.check_form_formular_ok = False

    def reformatting_sheet(self):
        try:
            self.wb = load_workbook(self.path)
            self.ws = self.wb.active

            self.list_merge_cells = copy(self.ws.merged_cells.ranges)
            if not self.list_merge_cells:
                raise Exception

            for item in self.list_merge_cells:
                self.bounds = item.bounds
                self.ws.unmerge_cells(str(item))
                self.copy_cells_ok = self.copy_cells(self.bounds)
                if not self.copy_cells_ok:
                    raise Exception

            self.merge_header_ok = self.merge_header(self.height_header, self.max_column)
            if not self.merge_header_ok:
                raise Exception
            self.check_form_formular_ok = self.check_form_formular()

            if self.check_form_formular_ok:
                return self.ws

            raise Exception

        except TypeError as _ex:
            print(f'[ERROR] -- "conversions/reformatting_sheet" -- Error type {_ex}')
        except Exception as _ex:
            print(f'[ERROR] -- "conversions/reformatting_sheet" -- {_ex}')

    def check_form_formular(self) -> bool:
        try:
            for i in FormularQ:
                if str.lower(self.ws.cell(row=1, column=i.value).value) != i.name:
                    raise TypeError
            return True
        except TypeError as _ex:
            print(f'[ERROR] --conversions/check_form_formular-- {_ex}')
        except Exception as _ex:
            print(f'[ERROR] --conversions/check_form_formular-- {_ex}')

    def copy_cells(self, bounds: tuple) -> bool:
        try:
            if bounds[1] == bounds[3] and \
                    bounds[0] != bounds[2] and bounds[0] < bounds[2]:  # row identical, copy columns

                for col in range(bounds[0] + 1, bounds[2] + 1):
                    self.ws.cell(row=bounds[1], column=col).value = \
                        self.ws.cell(row=bounds[1], column=bounds[0]).value
                return True

            elif bounds[0] == bounds[2] and \
                    bounds[1] != bounds[3] and bounds[1] < bounds[3]:  # column identical, copy row

                for row in range(bounds[1] + 1, bounds[3] + 1):
                    self.ws.cell(row=row, column=bounds[0]).value = \
                        self.ws.cell(row=bounds[1], column=bounds[0]).value
                return True

            else:
                print(f'Формат не распознан: {bounds}')
                raise Exception

        except TypeError as _ex:
            print(f'[ERROR] -- "conversions/copy_cells" -- Error type {_ex}')
            return False
        except Exception as _ex:
            print(f'[ERROR] -- "conversions/copy_cells" -- {_ex}')
            return False

    def merge_header(self, height_header: int, max_column: int) -> bool:
        try:
            for col in range(1, max_column + 1):
                list_cells_in_col = []
                for row in range(1, height_header + 1):
                    if self.ws.cell(row, col) is None or '' or self.ws.cell(row, col).value in list_cells_in_col:
                        continue
                    list_cells_in_col.append(str(self.ws.cell(row, col).value))

                if len(list_cells_in_col) > 1:
                    self.ws.cell(row=3, column=col).value = '_'.join(list_cells_in_col)
                else:
                    self.ws.cell(row=3, column=col).value = list_cells_in_col[0]

            self.ws.delete_rows(1, 2)
            return True
        except TypeError as _ex:
            print(f'[ERROR] -- "conversions/merge_header" -- Error type {_ex}')
            return False
        except Exception as _ex:
            print(f'[ERROR] -- "conversions/merge_header" -- {_ex}')
            return False
